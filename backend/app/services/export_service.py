from sqlalchemy.orm import Session
from datetime import date, datetime
from typing import Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO

from ..models.room import Room
from ..models.booking import Booking
from ..services.analytics_service import AnalyticsService


class ExportService:
    @staticmethod
    def export_rooms_to_excel(db: Session) -> bytes:
        """Export all rooms to Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Xonalar"

        # Headers
        headers = ["ID", "Xona raqami", "Xona turi", "Holati", "Yaratilgan"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)

        # Data
        rooms = db.query(Room).all()
        for row, room in enumerate(rooms, 2):
            ws.cell(row=row, column=1, value=room.id)
            ws.cell(row=row, column=2, value=room.room_number)
            ws.cell(row=row, column=3, value=room.room_type.value if room.room_type else "")
            ws.cell(row=row, column=4, value="Bo'sh" if room.is_available else "Band")
            ws.cell(row=row, column=5, value=room.created_at.strftime("%Y-%m-%d %H:%M"))

        # Auto-adjust columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    @staticmethod
    def export_bookings_to_excel(db: Session, start_date: Optional[date] = None,
                                 end_date: Optional[date] = None) -> bytes:
        """Export bookings to Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Bronlar"

        # Headers
        headers = ["ID", "Xona", "Kirish sanasi", "Chiqish sanasi", "Mehmon", "Izohlar", "Yaratilgan", "Yaratgan"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)

        # Query bookings
        query = db.query(Booking)
        if start_date:
            query = query.filter(Booking.end_date >= start_date)
        if end_date:
            query = query.filter(Booking.start_date <= end_date)

        bookings = query.order_by(Booking.start_date).all()

        # Data
        for row, booking in enumerate(bookings, 2):
            ws.cell(row=row, column=1, value=booking.id)
            ws.cell(row=row, column=2, value=f"№{booking.room.room_number}" if booking.room else "")
            ws.cell(row=row, column=3, value=booking.start_date.strftime("%Y-%m-%d"))
            ws.cell(row=row, column=4, value=booking.end_date.strftime("%Y-%m-%d"))
            ws.cell(row=row, column=5, value=booking.guest_name or "")
            ws.cell(row=row, column=6, value=booking.notes or "")
            ws.cell(row=row, column=7, value=booking.created_at.strftime("%Y-%m-%d %H:%M"))
            ws.cell(row=row, column=8, value=booking.user.full_name if booking.user else "")

        # Auto-adjust columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    @staticmethod
    def export_analytics_to_excel(db: Session, start_date: date, end_date: date) -> bytes:
        """Export analytics report to Excel"""
        wb = Workbook()

        # 1. Occupancy Stats
        ws1 = wb.active
        ws1.title = "Bandlik statistikasi"

        occupancy_stats = AnalyticsService.get_occupancy_stats(db, start_date, end_date)

        # Summary
        ws1.cell(row=1, column=1, value="Davr")
        ws1.cell(row=1, column=2, value=f"{start_date} - {end_date}")
        ws1.cell(row=2, column=1, value="O'rtacha bandlik")
        ws1.cell(row=2, column=2, value=f"{occupancy_stats['average_occupancy']}%")
        ws1.cell(row=3, column=1, value="Jami xonalar")
        ws1.cell(row=3, column=2, value=occupancy_stats['total_rooms'])

        # Daily stats
        ws1.cell(row=5, column=1, value="Sana")
        ws1.cell(row=5, column=2, value="Band")
        ws1.cell(row=5, column=3, value="Bo'sh")
        ws1.cell(row=5, column=4, value="Bandlik %")

        for i, daily in enumerate(occupancy_stats['daily_stats'], 6):
            ws1.cell(row=i, column=1, value=daily['date'])
            ws1.cell(row=i, column=2, value=daily['occupied'])
            ws1.cell(row=i, column=3, value=daily['available'])
            ws1.cell(row=i, column=4, value=f"{daily['occupancy_rate']}%")

        # 2. Room Type Stats
        ws2 = wb.create_sheet("Xona turlari")
        room_type_stats = AnalyticsService.get_room_type_stats(db, start_date, end_date)

        headers = ["Xona turi", "Jami xonalar", "Bronlar soni", "Band kunlar", "Bandlik %"]
        for col, header in enumerate(headers, 1):
            ws2.cell(row=1, column=col, value=header)

        for row, stat in enumerate(room_type_stats, 2):
            ws2.cell(row=row, column=1, value=stat['room_type'])
            ws2.cell(row=row, column=2, value=stat['total_rooms'])
            ws2.cell(row=row, column=3, value=stat['bookings_count'])
            ws2.cell(row=row, column=4, value=stat['total_booked_days'])
            ws2.cell(row=row, column=5, value=f"{stat['occupancy_rate']}%")

        # 3. Trends
        ws3 = wb.create_sheet("Tendensiyalar")
        trends = AnalyticsService.get_booking_trends(db, 6)

        ws3.cell(row=1, column=1, value="Oy")
        ws3.cell(row=1, column=2, value="Bronlar soni")
        ws3.cell(row=1, column=3, value="O'rtacha muddat")

        for row, trend in enumerate(trends, 2):
            ws3.cell(row=row, column=1, value=trend['month'])
            ws3.cell(row=row, column=2, value=trend['bookings_count'])
            ws3.cell(row=row, column=3, value=trend['average_duration'])

        # Format all sheets
        for ws in wb.worksheets:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column_letter].width = adjusted_width

        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()